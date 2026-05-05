import csv
from pathlib import Path

from openpyxl import load_workbook

from autotache_jobs.exporter import (
    CSV_COLUMNS,
    export_offers_to_csv,
    export_offers_to_tracking_xlsx,
    export_offers_to_xlsx,
)


def _sample_offer(**overrides) -> dict:
    values = {
        "date_detection": "2026-04-29T10:00:00",
        "id_offre": "A1",
        "source": "France Travail",
        "titre": "Developpeur WordPress",
        "entreprise": "Agence Numérique",
        "localisation": "Lyon",
        "code_postal": "69000",
        "type_contrat": "CDI",
        "experience": "2 ans",
        "salaire_brut": "Annuel de 35000 Euros a 42000 Euros",
        "salaire_min": 35000,
        "salaire_max": 42000,
        "salaire_moyen": 38500,
        "salaire_type": "annuel",
        "teletravail_mention": "teletravail partiel",
        "teletravail_jours": 2,
        "technologies": ["WordPress", "Figma"],
        "date_publication": "2026-04-20",
        "date_actualisation": "2026-04-21",
        "is_relevant": True,
        "decision": "Pertinent",
        "score_total": 87,
        "score_reason": "Pertinent: score 87/100.",
        "score_details": {
            "technologies": {"score": 35},
            "titre": {"score": 25},
            "domaine_metier": {"score": 12},
            "localisation": {"score": 10},
            "teletravail": {"score": 10},
            "contrat": {"score": 5},
            "penalties": {"score": 0},
            "eliminatory_reason": None,
        },
        "relevance_reason": "Offre conservée",
        "matched_keywords": ["WordPress", "Figma"],
        "excluded_by": [],
        "url_offre": "https://candidat.francetravail.fr/offres/recherche/detail/A1",
    }
    values.update(overrides)
    return values


def test_export_offers_to_csv_creates_file(tmp_path: Path) -> None:
    output = export_offers_to_csv([_sample_offer()], tmp_path, filename="offres.csv")

    assert output == tmp_path / "offres.csv"
    assert output.exists()


def test_export_offers_to_csv_empty_list_creates_nothing(tmp_path: Path) -> None:
    output = export_offers_to_csv([], tmp_path, filename="offres.csv")

    assert output is None
    assert not (tmp_path / "offres.csv").exists()


def test_export_csv_contains_expected_columns(tmp_path: Path) -> None:
    output = export_offers_to_csv([_sample_offer()], tmp_path, filename="offres.csv")

    with output.open("r", encoding="utf-8-sig", newline="") as csv_file:
        reader = csv.reader(csv_file, delimiter=";")
        header = next(reader)

    assert header == CSV_COLUMNS


def test_export_csv_encodes_accents_with_utf8_sig(tmp_path: Path) -> None:
    output = export_offers_to_csv([_sample_offer()], tmp_path, filename="offres.csv")
    raw_bytes = output.read_bytes()

    assert raw_bytes.startswith(b"\xef\xbb\xbf")
    assert "Agence Numérique" in raw_bytes.decode("utf-8-sig")


def test_export_csv_converts_lists_to_readable_text(tmp_path: Path) -> None:
    output = export_offers_to_csv([_sample_offer()], tmp_path, filename="offres.csv")

    with output.open("r", encoding="utf-8-sig", newline="") as csv_file:
        rows = list(csv.DictReader(csv_file, delimiter=";"))

    assert rows[0]["technologies"] == "WordPress, Figma"
    assert rows[0]["matched_keywords"] == "WordPress, Figma"
    assert rows[0]["excluded_by"] == ""


def test_export_csv_converts_score_details_to_readable_text(tmp_path: Path) -> None:
    output = export_offers_to_csv([_sample_offer()], tmp_path, filename="debug_offres.csv")

    with output.open("r", encoding="utf-8-sig", newline="") as csv_file:
        rows = list(csv.DictReader(csv_file, delimiter=";"))

    assert rows[0]["decision"] == "Pertinent"
    assert rows[0]["score_total"] == "87"
    assert rows[0]["score_reason"] == "Pertinent: score 87/100."
    assert "technologies=35" in rows[0]["score_details"]
    assert "titre=25" in rows[0]["score_details"]


def test_export_offers_to_xlsx_creates_file(tmp_path: Path) -> None:
    output = export_offers_to_xlsx([_sample_offer()], tmp_path, filename="offres.xlsx")

    assert output == tmp_path / "offres.xlsx"
    assert output.exists()


def test_export_offers_to_xlsx_empty_list_creates_nothing(tmp_path: Path) -> None:
    output = export_offers_to_xlsx([], tmp_path, filename="offres.xlsx")

    assert output is None
    assert not (tmp_path / "offres.xlsx").exists()


def test_export_xlsx_contains_expected_columns(tmp_path: Path) -> None:
    output = export_offers_to_xlsx([_sample_offer()], tmp_path, filename="offres.xlsx")
    workbook = load_workbook(output)
    worksheet = workbook["Offres"]

    header = [cell.value for cell in worksheet[1]]

    assert header == CSV_COLUMNS


def test_export_xlsx_converts_lists_to_readable_text(tmp_path: Path) -> None:
    output = export_offers_to_xlsx([_sample_offer()], tmp_path, filename="offres.xlsx")
    workbook = load_workbook(output)
    worksheet = workbook["Offres"]
    columns = {cell.value: cell.column for cell in worksheet[1]}

    assert worksheet.cell(row=2, column=columns["technologies"]).value == "WordPress, Figma"
    assert worksheet.cell(row=2, column=columns["matched_keywords"]).value == "WordPress, Figma"
    assert worksheet.cell(row=2, column=columns["excluded_by"]).value in {"", None}
    assert worksheet.cell(row=2, column=columns["is_relevant"]).value == "Oui"


def test_export_xlsx_converts_score_details_to_readable_text(tmp_path: Path) -> None:
    output = export_offers_to_xlsx([_sample_offer()], tmp_path, filename="debug_offres.xlsx")
    workbook = load_workbook(output)
    worksheet = workbook["Offres"]
    columns = {cell.value: cell.column for cell in worksheet[1]}

    assert worksheet.cell(row=2, column=columns["decision"]).value == "Pertinent"
    assert worksheet.cell(row=2, column=columns["score_total"]).value == 87
    assert worksheet.cell(row=2, column=columns["score_reason"]).value == "Pertinent: score 87/100."
    score_details = worksheet.cell(row=2, column=columns["score_details"]).value
    assert "technologies=35" in score_details
    assert "titre=25" in score_details


def test_export_xlsx_url_is_clickable(tmp_path: Path) -> None:
    output = export_offers_to_xlsx([_sample_offer()], tmp_path, filename="offres.xlsx")
    workbook = load_workbook(output)
    worksheet = workbook["Offres"]
    columns = {cell.value: cell.column for cell in worksheet[1]}
    url_cell = worksheet.cell(row=2, column=columns["url_offre"])

    assert url_cell.value == "https://candidat.francetravail.fr/offres/recherche/detail/A1"
    assert url_cell.hyperlink is not None
    assert url_cell.hyperlink.target == url_cell.value


def test_tracking_xlsx_is_created_when_missing(tmp_path: Path) -> None:
    output = export_offers_to_tracking_xlsx([_sample_offer()], tmp_path)

    assert output == tmp_path / "offres_suivi.xlsx"
    assert output.exists()

    workbook = load_workbook(output)
    worksheet = workbook["Offres"]

    assert [cell.value for cell in worksheet[1]] == CSV_COLUMNS
    assert worksheet.max_row == 2


def test_tracking_xlsx_appends_new_offers(tmp_path: Path) -> None:
    output = export_offers_to_tracking_xlsx([_sample_offer(id_offre="A1")], tmp_path)
    export_offers_to_tracking_xlsx([_sample_offer(id_offre="A2", url_offre="https://example.test/A2")], tmp_path)

    workbook = load_workbook(output)
    worksheet = workbook["Offres"]
    columns = {cell.value: cell.column for cell in worksheet[1]}

    assert worksheet.max_row == 3
    assert worksheet.cell(row=2, column=columns["id_offre"]).value == "A1"
    assert worksheet.cell(row=3, column=columns["id_offre"]).value == "A2"


def test_tracking_xlsx_does_not_add_same_source_id_twice(tmp_path: Path) -> None:
    offer = _sample_offer(id_offre="A1")

    output = export_offers_to_tracking_xlsx([offer], tmp_path)
    export_offers_to_tracking_xlsx([offer], tmp_path)

    workbook = load_workbook(output)
    worksheet = workbook["Offres"]

    assert worksheet.max_row == 2


def test_tracking_xlsx_ignores_rejected_offers(tmp_path: Path) -> None:
    output = export_offers_to_tracking_xlsx([_sample_offer(id_offre="A1", decision="Rejeté")], tmp_path)

    workbook = load_workbook(output)
    worksheet = workbook["Offres"]

    assert worksheet.max_row == 1


def test_tracking_xlsx_deduplicates_by_source_and_url_when_id_missing(tmp_path: Path) -> None:
    offer = _sample_offer(id_offre="", url_offre="https://example.test/job")

    output = export_offers_to_tracking_xlsx([offer], tmp_path)
    export_offers_to_tracking_xlsx([offer], tmp_path)

    workbook = load_workbook(output)
    worksheet = workbook["Offres"]

    assert worksheet.max_row == 2


def test_tracking_xlsx_deduplicates_by_source_title_and_company_when_id_and_url_missing(tmp_path: Path) -> None:
    offer = _sample_offer(id_offre="", url_offre="", titre="Designer web", entreprise="Studio")

    output = export_offers_to_tracking_xlsx([offer], tmp_path)
    export_offers_to_tracking_xlsx([offer], tmp_path)

    workbook = load_workbook(output)
    worksheet = workbook["Offres"]

    assert worksheet.max_row == 2
