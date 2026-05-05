"""Export helpers for normalized offers."""

from __future__ import annotations

import csv
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


CSV_COLUMNS = [
    "date_detection",
    "id_offre",
    "source",
    "titre",
    "entreprise",
    "localisation",
    "code_postal",
    "type_contrat",
    "experience",
    "salaire_brut",
    "salaire_min",
    "salaire_max",
    "salaire_moyen",
    "salaire_type",
    "teletravail_mention",
    "teletravail_jours",
    "technologies",
    "date_publication",
    "date_actualisation",
    "is_relevant",
    "relevance_reason",
    "decision",
    "score_total",
    "score_reason",
    "score_details",
    "matched_keywords",
    "excluded_by",
    "url_offre",
]

XLSX_RELEVANT_FILL = PatternFill(fill_type="solid", fgColor="EAF4EA")
XLSX_HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
XLSX_MAX_COLUMN_WIDTH = 60
TRACKING_XLSX_FILENAME = "offres_suivi.xlsx"
EXPORTABLE_DECISIONS = {"Pertinent", "À vérifier", "Ã€ vÃ©rifier"}


def export_offers_to_csv(
    offers: list[dict],
    export_dir: str | Path,
    filename: str | None = None,
) -> Path | None:
    """Export normalized offers to an Excel-friendly French CSV."""

    if not offers:
        return None

    target_dir = Path(export_dir)
    target_dir.mkdir(parents=True, exist_ok=True)

    if filename is None:
        filename = f"offres_{datetime.now().strftime('%Y-%m-%d_%H%M')}.csv"

    output_path = target_dir / filename
    with output_path.open("w", encoding="utf-8-sig", newline="") as csv_file:
        writer = csv.DictWriter(
            csv_file,
            fieldnames=CSV_COLUMNS,
            delimiter=";",
            extrasaction="ignore",
        )
        writer.writeheader()
        for offer in offers:
            writer.writerow({column: _format_csv_value(offer.get(column)) for column in CSV_COLUMNS})

    return output_path


def export_offers_to_xlsx(
    offers: list[dict],
    export_dir: str | Path,
    filename: str | None = None,
) -> Path | None:
    """Export normalized offers to a formatted Excel workbook."""

    if not offers:
        return None

    target_dir = Path(export_dir)
    target_dir.mkdir(parents=True, exist_ok=True)

    if filename is None:
        filename = f"offres_{datetime.now().strftime('%Y-%m-%d_%H%M')}.xlsx"

    output_path = target_dir / filename
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Offres"

    worksheet.append(CSV_COLUMNS)
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = XLSX_HEADER_FILL

    url_column_index = CSV_COLUMNS.index("url_offre") + 1
    is_relevant_column_index = CSV_COLUMNS.index("is_relevant") + 1

    for offer in offers:
        row_values = [_format_xlsx_value(column, offer.get(column)) for column in CSV_COLUMNS]
        worksheet.append(row_values)
        row_index = worksheet.max_row

        if offer.get("is_relevant") is True:
            for cell in worksheet[row_index]:
                cell.fill = XLSX_RELEVANT_FILL

        url_cell = worksheet.cell(row=row_index, column=url_column_index)
        if url_cell.value:
            url_cell.hyperlink = str(url_cell.value)
            url_cell.style = "Hyperlink"

        relevant_cell = worksheet.cell(row=row_index, column=is_relevant_column_index)
        relevant_cell.value = _format_relevance_value(offer.get("is_relevant"))

    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.freeze_panes = "A2"
    _adjust_column_widths(worksheet)

    workbook.save(output_path)
    return output_path


def export_offers_to_tracking_xlsx(
    offers: list[dict],
    export_dir: str | Path,
    filename: str = TRACKING_XLSX_FILENAME,
) -> Path:
    """Create or update the cumulative tracking workbook with exportable offers only."""

    target_dir = Path(export_dir)
    target_dir.mkdir(parents=True, exist_ok=True)
    output_path = target_dir / filename

    if output_path.exists():
        workbook = load_workbook(output_path)
        worksheet = workbook["Offres"] if "Offres" in workbook.sheetnames else workbook.active
        worksheet.title = "Offres"
        _ensure_header(worksheet)
    else:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Offres"
        worksheet.append(CSV_COLUMNS)

    existing_keys = _existing_tracking_keys(worksheet)
    for offer in offers:
        if not _is_tracking_exportable(offer):
            continue

        tracking_key = _tracking_key_from_offer(offer)
        if not tracking_key or tracking_key in existing_keys:
            continue

        worksheet.append([_format_xlsx_value(column, offer.get(column)) for column in CSV_COLUMNS])
        existing_keys.add(tracking_key)

    _format_tracking_worksheet(worksheet)
    workbook.save(output_path)
    return output_path


def _format_csv_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, list):
        return ", ".join(str(item) for item in value)
    if isinstance(value, dict):
        return _format_dict_value(value)
    return str(value)


def _format_xlsx_value(column: str, value: Any) -> Any:
    if column == "is_relevant":
        return _format_relevance_value(value)
    if value is None:
        return ""
    if isinstance(value, list):
        return ", ".join(str(item) for item in value)
    if isinstance(value, dict):
        return _format_dict_value(value)
    return value


def _format_relevance_value(value: Any) -> str:
    if value is True:
        return "Oui"
    if value is False:
        return "Non"
    return ""


def _ensure_header(worksheet: Any) -> None:
    header = [cell.value for cell in worksheet[1]]
    if header == CSV_COLUMNS:
        return

    if worksheet.max_row == 1 and all(value is None for value in header):
        for column_index, column_name in enumerate(CSV_COLUMNS, start=1):
            worksheet.cell(row=1, column=column_index).value = column_name
        return

    worksheet.insert_rows(1)
    for column_index, column_name in enumerate(CSV_COLUMNS, start=1):
        worksheet.cell(row=1, column=column_index).value = column_name


def _format_tracking_worksheet(worksheet: Any) -> None:
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = XLSX_HEADER_FILL

    url_column_index = CSV_COLUMNS.index("url_offre") + 1
    is_relevant_column_index = CSV_COLUMNS.index("is_relevant") + 1

    for row_index in range(2, worksheet.max_row + 1):
        relevant_cell = worksheet.cell(row=row_index, column=is_relevant_column_index)
        if relevant_cell.value is True:
            relevant_cell.value = "Oui"
        elif relevant_cell.value is False:
            relevant_cell.value = "Non"
        if relevant_cell.value == "Oui":
            for cell in worksheet[row_index]:
                cell.fill = XLSX_RELEVANT_FILL

        url_cell = worksheet.cell(row=row_index, column=url_column_index)
        if url_cell.value:
            url_cell.hyperlink = str(url_cell.value)
            url_cell.style = "Hyperlink"

    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.freeze_panes = "A2"
    _adjust_column_widths(worksheet)


def _existing_tracking_keys(worksheet: Any) -> set[tuple[str, str, str]]:
    header = [cell.value for cell in worksheet[1]]
    columns = {column_name: index + 1 for index, column_name in enumerate(header) if column_name}
    keys: set[tuple[str, str, str]] = set()

    for row_index in range(2, worksheet.max_row + 1):
        row_values = {
            column: worksheet.cell(row=row_index, column=column_index).value
            for column, column_index in columns.items()
        }
        key = _tracking_key_from_values(row_values)
        if key:
            keys.add(key)

    return keys


def _tracking_key_from_offer(offer: dict) -> tuple[str, str, str] | None:
    return _tracking_key_from_values(offer)


def _tracking_key_from_values(values: dict[str, Any]) -> tuple[str, str, str] | None:
    source = _clean_key_part(values.get("source"))
    if not source:
        return None

    offer_id = _clean_key_part(values.get("id_offre"))
    if offer_id:
        return ("id", source, offer_id)

    url = _clean_key_part(values.get("url_offre"))
    if url:
        return ("url", source, url)

    title = _clean_key_part(values.get("titre"))
    company = _clean_key_part(values.get("entreprise"))
    if title or company:
        return ("title_company", source, f"{title}|{company}")

    return None


def _clean_key_part(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().casefold()


def _is_tracking_exportable(offer: dict) -> bool:
    return offer.get("decision") in EXPORTABLE_DECISIONS


def _adjust_column_widths(worksheet: Any) -> None:
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            if cell.value is None:
                continue
            max_length = max(max_length, len(str(cell.value)))
        worksheet.column_dimensions[column_letter].width = min(max_length + 2, XLSX_MAX_COLUMN_WIDTH)


def _format_dict_value(value: dict[str, Any]) -> str:
    score_parts = []
    for key in ["technologies", "titre", "domaine_metier", "localisation", "teletravail", "contrat"]:
        detail = value.get(key)
        if isinstance(detail, dict):
            score_parts.append(f"{key}={detail.get('score', 0)}")

    penalties = value.get("penalties")
    if isinstance(penalties, dict) and penalties.get("score", 0):
        score_parts.append(f"penalites={penalties.get('score', 0)}")

    eliminatory_reason = value.get("eliminatory_reason")
    if eliminatory_reason:
        score_parts.append(f"eliminatoire={eliminatory_reason}")

    if score_parts:
        return "; ".join(score_parts)

    return "; ".join(f"{key}={item}" for key, item in value.items())
