from pathlib import Path

from openpyxl import Workbook, load_workbook

from autotache_jobs.cv.generator import SECTIONS, generate_cv_draft
from autotache_jobs.cv.offer_reader import read_offer_from_xlsx
from autotache_jobs.cv.profile import load_profile


def test_generates_draft_markdown_without_modifying_sources_by_default(tmp_path: Path) -> None:
    profile_path = _write_profile(tmp_path)
    xlsx_path = _write_xlsx(tmp_path)

    profile_before = profile_path.read_bytes()
    xlsx_before_rows = load_workbook(xlsx_path).active.max_row
    profile = load_profile(profile_path)
    offer = read_offer_from_xlsx(xlsx_path, 2)

    output_path = generate_cv_draft(offer=offer, profile=profile, output_dir=tmp_path / "generated")

    content = output_path.read_text(encoding="utf-8")
    assert output_path.parent == tmp_path / "generated"
    assert output_path.name == "cv_draft_agence_test_integrateur_front_react.md"
    assert "# CV ciblé" in content
    assert "## CV proposé" in content
    assert "## Analyse de correspondance" in content
    assert "### Expériences à valoriser" in content
    assert "### Projets à valoriser" in content
    assert "### Formation" in content
    assert "### Informations offre" in content
    assert "### Points à ne pas sur-vendre" in content
    assert "Agence Test" in content
    assert "Integrateur front React" in content
    for section in SECTIONS:
        assert f"## {section}" in content
    assert "Resume court issu de profile_summary." in content
    assert "À orienter vers l'offre avec" in content
    assert "#### Integrateur web - Studio Local (2021 -> 2024)" in content
    assert "Livraison de pages accessibles" in content
    assert "Refonte vitrine" in content
    assert "- Technologies: HTML, CSS" in content
    assert "Portfolio public test" in content
    assert "https://example.test/portfolio" in content
    assert "- Technologies: WordPress" in content
    assert "- DUT Informatique — IUT — Dijon — 2018 -> 2020" in content
    assert "#### Compétences confirmées pertinentes\n- HTML\n- CSS" in content
    assert "#### Compétences complémentaires\n- JavaScript" in content
    assert "#### Compétences à confirmer\n- React (à confirmer, ne pas présenter comme maîtrisé)" in content
    assert "React: à confirmer, ne pas présenter comme maîtrisé." in content
    assert "Ne jamais inventer une expérience." in content
    assert "Score: 90" in content
    assert "Decision AutoTache: Pertinent" in content
    assert "URL: https://example.test/job" in content
    assert profile_path.read_bytes() == profile_before
    assert load_workbook(xlsx_path).active.max_row == xlsx_before_rows


def test_generates_recruiter_markdown_without_internal_analysis(tmp_path: Path) -> None:
    profile_path = _write_profile(tmp_path)
    xlsx_path = _write_xlsx(tmp_path)
    profile = load_profile(profile_path)
    offer = read_offer_from_xlsx(xlsx_path, 2)

    output_path = generate_cv_draft(
        offer=offer,
        profile=profile,
        output_dir=tmp_path / "generated",
        mode="recruiter",
    )

    content = output_path.read_text(encoding="utf-8")
    assert output_path.name == "cv_recruiter_agence_test_integrateur_front_react.md"
    assert "# CV - Bastien Test" in content
    assert "Developpeur web | Dijon | bastien@example.test | 0102030405" in content
    assert "Resume court issu de profile_summary." in content
    assert "approche orientée utilisateur" in content
    assert "HTML, CSS et JavaScript" in content
    assert "À orienter vers l'offre avec" not in content
    assert "Aucune compétence complémentaire" not in content
    assert "Aucune compétence à confirmer" not in content
    assert "## Profil ciblé" in content
    assert "## Compétences clés" in content
    assert "## Expériences" in content
    assert "## Projets" in content
    assert "## Formation" in content
    assert "## Analyse de correspondance" not in content
    assert "Points à ne pas sur-vendre" not in content
    assert "Règles de prudence" not in content
    assert "Mots-clés détectés" not in content
    assert "Score: 90" not in content
    assert "Decision AutoTache" not in content
    assert "France Travail" not in content
    assert "https://example.test/job" not in content
    assert "React (à confirmer" not in content
    assert "Compétences à confirmer" not in content


def test_generate_cv_draft_rejects_invalid_mode(tmp_path: Path) -> None:
    profile = load_profile(_write_profile(tmp_path))
    offer = read_offer_from_xlsx(_write_xlsx(tmp_path), 2)

    try:
        generate_cv_draft(offer=offer, profile=profile, output_dir=tmp_path / "generated", mode="invalid")
    except ValueError as exc:
        assert "--mode doit etre 'draft' ou 'recruiter'." in str(exc)
    else:
        raise AssertionError("generate_cv_draft should reject invalid modes")


def _write_profile(tmp_path: Path) -> Path:
    profile_path = tmp_path / "profile.yaml"
    profile_path.write_text(
        """
profile_summary:
  short: Resume court issu de profile_summary.
identity:
  name: Bastien Test
  title: Developpeur web
  location: Dijon
  email: bastien@example.test
  phone: "0102030405"
competences_fortes:
  - HTML
  - CSS
competences_moyennes:
  - JavaScript
to_confirm:
  - React
experiences:
  - title: Integrateur web
    company: Studio Local
    period: 2021 -> 2024
    summary: Integration responsive documentee dans le profil.
    achievements:
      - Livraison de pages accessibles
    projects:
      - name: Refonte vitrine
        bullets:
          - Composants HTML CSS
        technologies:
          - HTML
          - CSS
education:
  - degree: DUT Informatique
    school: IUT
    location: Dijon
    period: 2018 -> 2020
portfolio:
  public_projects:
    - name: Portfolio public test
      url: https://example.test/portfolio
      highlights:
        - Interface responsive
      technologies:
        - WordPress
""",
        encoding="utf-8",
    )
    return profile_path


def _write_xlsx(tmp_path: Path) -> Path:
    xlsx_path = tmp_path / "debug.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(
        [
            "source",
            "titre",
            "entreprise",
            "localisation",
            "description",
            "technologies",
            "type_contrat",
            "decision",
            "score_total",
            "url_offre",
        ]
    )
    worksheet.append(
        [
            "France Travail",
            "Integrateur front React",
            "Agence Test",
            "Dijon",
            "HTML CSS JavaScript React",
            "HTML, CSS, JavaScript, React",
            "CDI",
            "Pertinent",
            90,
            "https://example.test/job",
        ]
    )
    workbook.save(xlsx_path)
    return xlsx_path
