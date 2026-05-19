import csv
import json
import re
from pathlib import Path

import httpx
from openpyxl import load_workbook

from autotache_jobs.models import AppConfig, FranceTravailEnv
import autotache_jobs.runner as runner_module
from autotache_jobs.runner import run_job_search
from autotache_jobs.scoring import DECISION_RELEVANT, DECISION_REVIEW


class FakeFranceTravailClient:
    def __init__(self, results_by_call: list[list[dict]]) -> None:
        self.results_by_call = results_by_call
        self.calls: list[dict] = []

    def search_offers(self, **kwargs) -> list[dict]:
        self.calls.append(kwargs)
        if not self.results_by_call:
            return []
        return self.results_by_call.pop(0)


def _config(**overrides) -> AppConfig:
    values = {
        "keywords": ["wordpress"],
        "communes": ["75056"],
        "distance_km": 20,
        "contract_types": ["CDI"],
        "days_back": 7,
        "allow_internship": False,
        "allow_apprenticeship": False,
    }
    values.update(overrides)
    return AppConfig(**values)


def _env(**overrides) -> FranceTravailEnv:
    values = {
        "client_id": "fake-client-id",
        "client_secret": "fake-client-secret",
        "scope": "fake-scope",
        "token_url": "https://example.test/token",
        "api_base_url": "https://example.test/api",
        "adzuna_app_id": "fake-adzuna-id",
        "adzuna_app_key": "fake-adzuna-key",
        "jooble_api_key": "fake-jooble-key",
    }
    values.update(overrides)
    return FranceTravailEnv(
        **values,
    )


def _wordpress_offer(offer_id: str = "A1") -> dict:
    return {
        "id": offer_id,
        "intitule": "Developpeur WordPress front-end",
        "description": (
            "Creation de sites WordPress et WooCommerce avec Elementor. "
            "Integration web HTML CSS JavaScript, interface responsive et UI/UX. "
            "Teletravail partiel 2 jours."
        ),
        "typeContratLibelle": "CDI",
        "lieuTravail": {"libelle": "Dijon 21000", "codePostal": "21000"},
        "competences": [
            {"libelle": "WordPress"},
            {"libelle": "Elementor"},
            {"libelle": "WooCommerce"},
            {"libelle": "HTML"},
            {"libelle": "CSS"},
            {"libelle": "JavaScript"},
        ],
    }


def _arbeitnow_offer(offer_id: str = "arbeitnow-front") -> dict:
    return {
        "slug": offer_id,
        "title": "Frontend Developer WordPress",
        "description": (
            "Build WordPress interfaces with Elementor, HTML, CSS and JavaScript. "
            "Remote work possible."
        ),
        "company_name": "Remote Studio",
        "location": "Remote",
        "remote": True,
        "job_types": ["Full-time"],
        "tags": ["WordPress", "Elementor"],
        "url": f"https://www.arbeitnow.com/jobs/{offer_id}",
    }


def _remotive_offer(offer_id: str = "remotive-front") -> dict:
    return {
        "id": offer_id,
        "title": "Frontend Developer WordPress",
        "description": (
            "Build WordPress interfaces with Elementor, HTML, CSS and JavaScript. "
            "Remote work possible."
        ),
        "company_name": "Remote Studio",
        "candidate_required_location": "Worldwide",
        "job_type": "full_time",
        "category": "Software Development",
        "tags": ["WordPress", "Elementor"],
        "url": f"https://remotive.com/remote-jobs/software-dev/{offer_id}",
    }


def _adzuna_offer(offer_id: str = "adzuna-front") -> dict:
    return {
        "id": offer_id,
        "title": "Frontend Developer WordPress",
        "description": (
            "Build WordPress interfaces with Elementor, HTML, CSS and JavaScript. "
            "Remote work possible."
        ),
        "company": {"display_name": "Remote Studio"},
        "location": {"display_name": "Paris, France", "area": ["France", "Paris"]},
        "category": {"label": "IT Jobs"},
        "contract_type": "permanent",
        "redirect_url": f"https://www.adzuna.fr/details/{offer_id}",
    }


def _jooble_offer(offer_id: str = "jooble-front") -> dict:
    return {
        "id": offer_id,
        "title": "Frontend Developer WordPress",
        "snippet": (
            "Build WordPress interfaces with Elementor, HTML, CSS and JavaScript. "
            "Remote work possible."
        ),
        "company": "Remote Studio",
        "location": "Paris, France",
        "type": "CDI",
        "link": f"https://jooble.org/jdp/{offer_id}",
    }


def _arbeitnow_client(raw_offers: list[dict]) -> httpx.Client:
    def handler(request: httpx.Request) -> httpx.Response:
        return httpx.Response(200, json={"data": raw_offers, "links": {"next": None}, "meta": {}})

    return httpx.Client(transport=httpx.MockTransport(handler))


def _remotive_client(raw_offers: list[dict], calls: list[str] | None = None) -> httpx.Client:
    def handler(request: httpx.Request) -> httpx.Response:
        if calls is not None:
            calls.append(str(request.url))
        return httpx.Response(200, json={"jobs": raw_offers})

    return httpx.Client(transport=httpx.MockTransport(handler))


def _adzuna_client(raw_offers: list[dict], calls: list[str] | None = None) -> httpx.Client:
    def handler(request: httpx.Request) -> httpx.Response:
        if calls is not None:
            calls.append(str(request.url))
        return httpx.Response(200, json={"results": raw_offers})

    return httpx.Client(transport=httpx.MockTransport(handler))


def _jooble_client(raw_offers: list[dict], calls: list[str] | None = None) -> httpx.Client:
    def handler(request: httpx.Request) -> httpx.Response:
        if calls is not None:
            calls.append(str(request.url))
        return httpx.Response(200, json={"jobs": raw_offers})

    return httpx.Client(transport=httpx.MockTransport(handler))


def _themuse_offer(offer_id: str = "themuse-front") -> dict:
    return {
        "id": offer_id,
        "name": "Frontend Developer WordPress",
        "contents": (
            "Build WordPress interfaces with Elementor, HTML, CSS and JavaScript. "
            "Remote work possible."
        ),
        "company": {"name": "Remote Studio"},
        "locations": [{"name": "Paris, France"}],
        "categories": [{"name": "Software Engineering"}],
        "levels": [{"name": "Mid Level"}],
        "refs": {"landing_page": f"https://www.themuse.com/jobs/example/{offer_id}"},
    }


def _themuse_client(raw_offers: list[dict], calls: list[str] | None = None) -> httpx.Client:
    def handler(request: httpx.Request) -> httpx.Response:
        if calls is not None:
            calls.append(str(request.url))
        return httpx.Response(200, json={"page": 1, "page_count": 1, "results": raw_offers})

    return httpx.Client(transport=httpx.MockTransport(handler))


def _fake_scoring(decisions_by_offer_id: dict[str, str]):
    def fake_score(offer: dict) -> dict:
        decision = decisions_by_offer_id[str(offer["id_offre"])]
        return {
            "score_total": 90 if decision == DECISION_RELEVANT else 50,
            "decision": decision,
            "score_reason": f"decision forcee: {decision}",
            "score_details": {"test": decision},
        }

    return fake_score


def test_runner_exports_new_relevant_offer(tmp_path: Path) -> None:
    client = FakeFranceTravailClient([[_wordpress_offer()]])

    summary = run_job_search(_config(), _env(), client=client, data_dir=tmp_path / "data", export_dir=tmp_path / "exports")

    assert summary["total_raw"] == 1
    assert summary["total_normalized"] == 1
    assert summary["total_unique_normalized"] == 1
    assert summary["total_relevant"] == 1
    assert summary["total_new"] == 1
    assert summary["export_path"] is not None
    assert summary["xlsx_export_path"] is not None
    assert summary["tracking_xlsx_export_path"] is not None
    assert summary["debug_export_path"] is None
    assert summary["debug_xlsx_export_path"] is None
    assert Path(summary["export_path"]).exists()
    assert Path(summary["xlsx_export_path"]).exists()
    assert Path(summary["tracking_xlsx_export_path"]).exists()
    assert Path(summary["export_path"]).parent == tmp_path / "exports" / "offres"
    assert Path(summary["xlsx_export_path"]).parent == tmp_path / "exports" / "offres"
    assert Path(summary["tracking_xlsx_export_path"]).name == "offres_suivi.xlsx"
    assert summary["discord_enabled"] is False
    assert summary["discord_sent"] is False
    assert summary["discord_status"] == "disabled"


def test_runner_ignores_non_relevant_offer(tmp_path: Path) -> None:
    client = FakeFranceTravailClient(
        [[{"id": "DATA", "intitule": "Data analyst", "description": "Reporting BI et SQL."}]]
    )

    summary = run_job_search(_config(), _env(), client=client, data_dir=tmp_path / "data", export_dir=tmp_path / "exports")

    assert summary["total_raw"] == 1
    assert summary["total_relevant"] == 0
    assert summary["total_new"] == 0
    assert summary["export_path"] is None
    assert summary["xlsx_export_path"] is None
    assert summary["debug_export_path"] is None
    assert summary["debug_xlsx_export_path"] is None


def test_runner_ignores_already_seen_offer(tmp_path: Path) -> None:
    seen_path = tmp_path / "data" / "seen_offer_ids.json"
    seen_path.parent.mkdir(parents=True)
    seen_path.write_text(json.dumps(["A1"]), encoding="utf-8")
    client = FakeFranceTravailClient([[_wordpress_offer()]])

    summary = run_job_search(_config(), _env(), client=client, data_dir=tmp_path / "data", export_dir=tmp_path / "exports")

    assert summary["total_relevant"] == 1
    assert summary["total_new"] == 0
    assert summary["export_path"] is None
    assert summary["xlsx_export_path"] is None
    assert json.loads(seen_path.read_text(encoding="utf-8")) == ["A1"]


def test_runner_deduplicates_between_multiple_searches(tmp_path: Path) -> None:
    client = FakeFranceTravailClient([[_wordpress_offer("A1")], [_wordpress_offer("A1")]])
    config = _config(keywords=["wordpress", "elementor"])

    summary = run_job_search(
        config,
        _env(),
        client=client,
        data_dir=tmp_path / "data",
        export_dir=tmp_path / "exports",
        sleep_func=lambda _: None,
    )

    assert len(client.calls) == 2
    assert summary["total_raw"] == 2
    assert summary["total_normalized"] == 2
    assert summary["total_unique_normalized"] == 1
    assert summary["total_relevant"] == 1
    assert summary["total_new"] == 1


def test_runner_waits_between_api_calls(tmp_path: Path) -> None:
    sleep_calls = []
    client = FakeFranceTravailClient([[], [], []])
    config = _config(keywords=["wordpress", "elementor", "figma"])

    run_job_search(
        config,
        _env(),
        client=client,
        data_dir=tmp_path / "data",
        export_dir=tmp_path / "exports",
        sleep_func=sleep_calls.append,
    )

    assert len(client.calls) == 3
    assert sleep_calls == [0.8, 0.8]


def test_runner_returns_none_export_path_when_no_new_offer(tmp_path: Path) -> None:
    client = FakeFranceTravailClient([[]])

    summary = run_job_search(_config(), _env(), client=client, data_dir=tmp_path / "data", export_dir=tmp_path / "exports")

    assert summary["total_raw"] == 0
    assert summary["total_new"] == 0
    assert summary["export_path"] is None
    assert summary["xlsx_export_path"] is None


def test_runner_updates_seen_ids_only_with_exported_offers(tmp_path: Path) -> None:
    client = FakeFranceTravailClient(
        [
            [
                _wordpress_offer("A1"),
                {"id": "DATA", "intitule": "Data analyst", "description": "Reporting BI et SQL."},
            ]
        ]
    )

    summary = run_job_search(_config(), _env(), client=client, data_dir=tmp_path / "data", export_dir=tmp_path / "exports")
    seen_ids = json.loads(Path(summary["seen_ids_path"]).read_text(encoding="utf-8"))

    assert seen_ids == ["A1"]


def test_runner_uses_injected_client_without_network_client(tmp_path: Path) -> None:
    client = FakeFranceTravailClient([[_wordpress_offer()]])

    summary = run_job_search(_config(), _env(), client=client, data_dir=tmp_path / "data", export_dir=tmp_path / "exports")

    assert client.calls
    assert client.calls[0]["keyword"] == "wordpress"
    assert client.calls[0]["commune"] == "75056"
    assert client.calls[0]["distance"] == 20
    assert client.calls[0]["type_contrat"] == "CDI"
    assert re.fullmatch(r"\d{4}-\d{2}-\d{2}T00:00:00Z", client.calls[0]["min_creation_date"])
    assert re.fullmatch(r"\d{4}-\d{2}-\d{2}T23:59:59Z", client.calls[0]["max_creation_date"])
    assert summary["total_new"] == 1


def test_runner_can_include_debug_offers(tmp_path: Path) -> None:
    client = FakeFranceTravailClient(
        [[{"id": "DATA", "intitule": "Data analyst", "description": "Reporting BI et SQL."}]]
    )

    summary = run_job_search(
        _config(),
        _env(),
        client=client,
        data_dir=tmp_path / "data",
        export_dir=tmp_path / "exports",
        include_debug_offers=True,
    )

    assert summary["total_relevant"] == 0
    assert summary["export_path"] is None
    assert summary["xlsx_export_path"] is None
    assert summary["debug_export_path"] is not None
    assert summary["debug_xlsx_export_path"] is not None
    assert Path(summary["debug_export_path"]).exists()
    assert Path(summary["debug_xlsx_export_path"]).exists()
    assert Path(summary["debug_export_path"]).parent == tmp_path / "exports" / "debug"
    assert Path(summary["debug_xlsx_export_path"]).parent == tmp_path / "exports" / "debug"
    assert Path(summary["debug_export_path"]).name.startswith("debug_offres_")
    assert Path(summary["debug_xlsx_export_path"]).name.startswith("debug_offres_")
    assert summary["debug_offers"][0]["id_offre"] == "DATA"
    assert summary["debug_offers"][0]["is_relevant"] is False
    assert "Offre rejetee" in summary["debug_offers"][0]["relevance_reason"]
    assert summary["debug_offers"][0]["decision"] == "Rejeté"
    assert isinstance(summary["debug_offers"][0]["score_total"], int)
    assert summary["debug_offers"][0]["score_reason"]


def test_runner_debug_offers_are_deduplicated(tmp_path: Path) -> None:
    client = FakeFranceTravailClient([[_wordpress_offer("A1")], [_wordpress_offer("A1")]])
    config = _config(keywords=["wordpress", "elementor"])

    summary = run_job_search(
        config,
        _env(),
        client=client,
        data_dir=tmp_path / "data",
        export_dir=tmp_path / "exports",
        include_debug_offers=True,
        sleep_func=lambda _: None,
    )

    assert summary["total_normalized"] == 2
    assert summary["total_unique_normalized"] == 1
    assert [offer["id_offre"] for offer in summary["debug_offers"]] == ["A1"]
    assert summary["total_new"] == 1


def test_runner_debug_export_contains_unique_normalized_offers(tmp_path: Path) -> None:
    client = FakeFranceTravailClient(
        [
            [_wordpress_offer("A1")],
            [
                _wordpress_offer("A1"),
                {"id": "DATA", "intitule": "Data analyst", "description": "Reporting BI et SQL."},
            ],
        ]
    )
    config = _config(keywords=["wordpress", "elementor"])

    summary = run_job_search(
        config,
        _env(),
        client=client,
        data_dir=tmp_path / "data",
        export_dir=tmp_path / "exports",
        include_debug_offers=True,
        sleep_func=lambda _: None,
    )

    assert summary["total_normalized"] == 3
    assert summary["total_unique_normalized"] == 2
    assert summary["debug_export_path"] is not None
    with Path(summary["debug_export_path"]).open("r", encoding="utf-8-sig", newline="") as csv_file:
        rows = list(csv.DictReader(csv_file, delimiter=";"))

    assert [row["id_offre"] for row in rows] == ["A1", "DATA"]
    assert rows[0]["is_relevant"] == "True"
    assert rows[1]["is_relevant"] == "False"
    assert rows[0]["decision"] == "Pertinent"
    assert rows[0]["score_total"]
    assert rows[0]["score_reason"]
    assert "technologies=" in rows[0]["score_details"]


def test_runner_debug_xlsx_export_contains_scoring_columns(tmp_path: Path) -> None:
    client = FakeFranceTravailClient([[_wordpress_offer("A1")]])

    summary = run_job_search(
        _config(),
        _env(),
        client=client,
        data_dir=tmp_path / "data",
        export_dir=tmp_path / "exports",
        include_debug_offers=True,
    )

    workbook = load_workbook(summary["debug_xlsx_export_path"])
    worksheet = workbook["Offres"]
    columns = {cell.value: cell.column for cell in worksheet[1]}

    for column in ["decision", "score_total", "score_reason", "score_details"]:
        assert column in columns
    assert worksheet.cell(row=2, column=columns["decision"]).value == "Pertinent"
    assert isinstance(worksheet.cell(row=2, column=columns["score_total"]).value, int)
    assert worksheet.cell(row=2, column=columns["score_reason"]).value
    assert "technologies=" in worksheet.cell(row=2, column=columns["score_details"]).value


def test_runner_adds_scoring_fields_to_debug_offers(tmp_path: Path) -> None:
    client = FakeFranceTravailClient([[_wordpress_offer()]])

    summary = run_job_search(
        _config(),
        _env(),
        client=client,
        data_dir=tmp_path / "data",
        export_dir=tmp_path / "exports",
        include_debug_offers=True,
    )
    offer = summary["debug_offers"][0]

    assert "score_total" in offer
    assert "decision" in offer
    assert "score_reason" in offer
    assert "score_details" in offer
    assert offer["decision"] == "Pertinent"
    assert isinstance(offer["score_details"], dict)


def test_runner_returns_decision_counts_for_unique_debug_offers(tmp_path: Path) -> None:
    client = FakeFranceTravailClient(
        [
            [
                _wordpress_offer("A1"),
                {"id": "DATA", "intitule": "Data analyst", "description": "Reporting BI et SQL."},
            ]
        ]
    )

    summary = run_job_search(
        _config(),
        _env(),
        client=client,
        data_dir=tmp_path / "data",
        export_dir=tmp_path / "exports",
        include_debug_offers=True,
    )

    assert summary["decision_counts"]["Pertinent"] == 1
    assert summary["decision_counts"]["À vérifier"] == 0
    assert summary["decision_counts"]["Rejeté"] == 1


def test_runner_scoring_does_not_change_existing_relevance_logic(tmp_path: Path) -> None:
    client = FakeFranceTravailClient(
        [[{"id": "DATA", "intitule": "Data analyst", "description": "Reporting BI et SQL."}]]
    )

    summary = run_job_search(
        _config(),
        _env(),
        client=client,
        data_dir=tmp_path / "data",
        export_dir=tmp_path / "exports",
        include_debug_offers=True,
    )

    assert summary["debug_offers"][0]["is_relevant"] is False
    assert summary["debug_offers"][0]["decision"] == "Rejeté"
    assert summary["total_relevant"] == 0
    assert summary["total_new"] == 0
    assert summary["export_path"] is None
    assert summary["xlsx_export_path"] is None


def test_runner_discord_disabled_sends_nothing(tmp_path: Path) -> None:
    client = FakeFranceTravailClient([[_wordpress_offer()]])
    calls = []

    summary = run_job_search(
        _config(notifications={"discord_enabled": False}),
        _env(discord_webhook_url="https://discord.test/webhook"),
        client=client,
        data_dir=tmp_path / "data",
        export_dir=tmp_path / "exports",
        discord_sender=lambda *args: calls.append(args),
    )

    assert calls == []
    assert summary["discord_enabled"] is False
    assert summary["discord_sent"] is False
    assert summary["discord_status"] == "disabled"


def test_runner_discord_enabled_without_webhook_does_not_crash(tmp_path: Path) -> None:
    client = FakeFranceTravailClient([[_wordpress_offer()]])

    summary = run_job_search(
        _config(notifications={"discord_enabled": True}),
        _env(discord_webhook_url=""),
        client=client,
        data_dir=tmp_path / "data",
        export_dir=tmp_path / "exports",
    )

    assert summary["discord_enabled"] is True
    assert summary["discord_sent"] is False
    assert summary["discord_status"] == "webhook_missing"
    assert "DISCORD_WEBHOOK_URL" in summary["discord_error"]


def test_runner_discord_enabled_with_new_offers_sends_summary(tmp_path: Path) -> None:
    client = FakeFranceTravailClient([[_wordpress_offer()]])
    calls = []

    def fake_sender(webhook_url: str, summary: dict) -> dict:
        calls.append((webhook_url, summary.copy()))
        return {"sent": True, "status": "sent", "error": None}

    summary = run_job_search(
        _config(notifications={"discord_enabled": True}),
        _env(discord_webhook_url="https://discord.test/webhook"),
        client=client,
        data_dir=tmp_path / "data",
        export_dir=tmp_path / "exports",
        discord_sender=fake_sender,
    )

    assert len(calls) == 1
    assert calls[0][0] == "https://discord.test/webhook"
    assert calls[0][1]["total_new"] == 1
    assert summary["discord_sent"] is True
    assert summary["discord_status"] == "sent"


def test_runner_discord_no_relevant_offer_skips_by_default(tmp_path: Path) -> None:
    client = FakeFranceTravailClient(
        [[{"id": "DATA", "intitule": "Data analyst", "description": "Reporting BI et SQL."}]]
    )
    calls = []

    summary = run_job_search(
        _config(notifications={"discord_enabled": True, "notify_when_no_results": False}),
        _env(discord_webhook_url="https://discord.test/webhook"),
        client=client,
        data_dir=tmp_path / "data",
        export_dir=tmp_path / "exports",
        discord_sender=lambda *args: calls.append(args),
    )

    assert calls == []
    assert summary["discord_sent"] is False
    assert summary["discord_status"] == "skipped_no_relevant_offers"


def test_runner_discord_no_relevant_offer_sends_when_configured(tmp_path: Path) -> None:
    client = FakeFranceTravailClient(
        [[{"id": "DATA", "intitule": "Data analyst", "description": "Reporting BI et SQL."}]]
    )
    calls = []

    def fake_sender(webhook_url: str, summary: dict) -> dict:
        calls.append((webhook_url, summary.copy()))
        return {"sent": True, "status": "sent", "error": None}

    summary = run_job_search(
        _config(notifications={"discord_enabled": True, "notify_when_no_results": True}),
        _env(discord_webhook_url="https://discord.test/webhook"),
        client=client,
        data_dir=tmp_path / "data",
        export_dir=tmp_path / "exports",
        discord_sender=fake_sender,
    )

    assert len(calls) == 1
    assert calls[0][1]["total_relevant"] == 0
    assert summary["discord_sent"] is True


def test_runner_default_uses_france_travail_and_no_optional_sources(tmp_path: Path) -> None:
    client = FakeFranceTravailClient([[_wordpress_offer()]])
    remotive_calls = []
    adzuna_calls = []
    jooble_calls = []

    summary = run_job_search(
        _config(),
        _env(),
        client=client,
        arbeitnow_client=_arbeitnow_client([_arbeitnow_offer()]),
        remotive_client=_remotive_client([_remotive_offer()], calls=remotive_calls),
        adzuna_client=_adzuna_client([_adzuna_offer()], calls=adzuna_calls),
        jooble_client=_jooble_client([_jooble_offer()], calls=jooble_calls),
        data_dir=tmp_path / "data",
        export_dir=tmp_path / "exports",
    )

    assert len(client.calls) == 1
    assert remotive_calls == []
    assert adzuna_calls == []
    assert jooble_calls == []
    assert summary["total_raw"] == 1
    assert summary["sources_enabled"] == ["France Travail"]
    assert summary["source_counts"] == {"France Travail": {"raw": 1, "normalized": 1}}
    assert summary["source_stats"]["France Travail"] == {
        "enabled": True,
        "fetched": 1,
        "kept": 1,
        "filtered": 0,
    }
    assert summary["source_stats"]["Arbeitnow"] == {
        "enabled": False,
        "fetched": 0,
        "kept": 0,
        "filtered": 0,
    }
    assert summary["source_stats"]["Remotive"] == {
        "enabled": False,
        "fetched": 0,
        "kept": 0,
        "filtered": 0,
    }
    assert summary["source_stats"]["Adzuna"] == {
        "enabled": False,
        "fetched": 0,
        "kept": 0,
        "filtered": 0,
    }
    assert summary["source_stats"]["Jooble"] == {
        "enabled": False,
        "fetched": 0,
        "kept": 0,
        "filtered": 0,
    }


def test_runner_can_use_arbeitnow_when_france_travail_disabled(tmp_path: Path) -> None:
    client = FakeFranceTravailClient([[_wordpress_offer()]])

    summary = run_job_search(
        _config(sources={"france_travail": {"enabled": False}, "arbeitnow": {"enabled": True, "max_pages": 1}}),
        _env(),
        client=client,
        arbeitnow_client=_arbeitnow_client([_arbeitnow_offer()]),
        data_dir=tmp_path / "data",
        export_dir=tmp_path / "exports",
    )

    assert client.calls == []
    assert summary["total_raw"] == 1
    assert summary["total_relevant"] == 1
    assert summary["sources_enabled"] == ["Arbeitnow"]
    assert summary["source_counts"] == {"Arbeitnow": {"raw": 1, "normalized": 1}}
    assert summary["source_stats"]["France Travail"]["enabled"] is False
    assert summary["source_stats"]["Arbeitnow"] == {
        "enabled": True,
        "fetched": 1,
        "kept": 1,
        "filtered": 0,
    }


def test_runner_merges_france_travail_and_arbeitnow_sources(tmp_path: Path) -> None:
    client = FakeFranceTravailClient([[_wordpress_offer("FT1")]])

    summary = run_job_search(
        _config(sources={"france_travail": {"enabled": True}, "arbeitnow": {"enabled": True, "max_pages": 1}}),
        _env(),
        client=client,
        arbeitnow_client=_arbeitnow_client([_arbeitnow_offer("AN1")]),
        data_dir=tmp_path / "data",
        export_dir=tmp_path / "exports",
        include_debug_offers=True,
    )

    assert len(client.calls) == 1
    assert summary["total_raw"] == 2
    assert summary["total_unique_normalized"] == 2
    assert summary["total_relevant"] == 2
    assert summary["sources_enabled"] == ["France Travail", "Arbeitnow"]
    assert summary["source_counts"] == {
        "France Travail": {"raw": 1, "normalized": 1},
        "Arbeitnow": {"raw": 1, "normalized": 1},
    }
    assert summary["source_stats"]["France Travail"] == {
        "enabled": True,
        "fetched": 1,
        "kept": 1,
        "filtered": 0,
    }
    assert summary["source_stats"]["Arbeitnow"] == {
        "enabled": True,
        "fetched": 1,
        "kept": 1,
        "filtered": 0,
    }
    assert {offer["source"] for offer in summary["debug_offers"]} == {"France Travail", "Arbeitnow"}


def test_runner_can_use_remotive_when_other_sources_disabled(tmp_path: Path) -> None:
    client = FakeFranceTravailClient([[_wordpress_offer()]])
    remotive_calls = []

    summary = run_job_search(
        _config(sources={"france_travail": {"enabled": False}, "remotive": {"enabled": True, "keywords": ["frontend"]}}),
        _env(),
        client=client,
        remotive_client=_remotive_client([_remotive_offer("RM1")], calls=remotive_calls),
        data_dir=tmp_path / "data",
        export_dir=tmp_path / "exports",
    )

    assert client.calls == []
    assert remotive_calls == ["https://remotive.com/api/remote-jobs"]
    assert summary["total_raw"] == 1
    assert summary["total_relevant"] == 1
    assert summary["sources_enabled"] == ["Remotive"]
    assert summary["source_counts"] == {"Remotive": {"raw": 1, "normalized": 1}}
    assert summary["source_stats"]["France Travail"]["enabled"] is False
    assert summary["source_stats"]["Arbeitnow"]["enabled"] is False
    assert summary["source_stats"]["Remotive"] == {
        "enabled": True,
        "fetched": 1,
        "kept": 1,
        "filtered": 0,
    }


def test_runner_merges_france_travail_arbeitnow_and_remotive_sources(tmp_path: Path) -> None:
    client = FakeFranceTravailClient([[_wordpress_offer("FT1")]])

    summary = run_job_search(
        _config(
            sources={
                "france_travail": {"enabled": True},
                "arbeitnow": {"enabled": True, "max_pages": 1},
                "remotive": {"enabled": True, "keywords": ["frontend"]},
            }
        ),
        _env(),
        client=client,
        arbeitnow_client=_arbeitnow_client([_arbeitnow_offer("AN1")]),
        remotive_client=_remotive_client([_remotive_offer("RM1")]),
        data_dir=tmp_path / "data",
        export_dir=tmp_path / "exports",
        include_debug_offers=True,
    )

    assert len(client.calls) == 1
    assert summary["total_raw"] == 3
    assert summary["total_unique_normalized"] == 3
    assert summary["total_relevant"] == 3
    assert summary["sources_enabled"] == ["France Travail", "Arbeitnow", "Remotive"]
    assert summary["source_counts"] == {
        "France Travail": {"raw": 1, "normalized": 1},
        "Arbeitnow": {"raw": 1, "normalized": 1},
        "Remotive": {"raw": 1, "normalized": 1},
    }
    assert summary["source_stats"]["Remotive"] == {
        "enabled": True,
        "fetched": 1,
        "kept": 1,
        "filtered": 0,
    }
    assert {offer["source"] for offer in summary["debug_offers"]} == {"France Travail", "Arbeitnow", "Remotive"}


def test_runner_can_use_adzuna_when_other_sources_disabled(tmp_path: Path) -> None:
    client = FakeFranceTravailClient([[_wordpress_offer()]])
    adzuna_calls = []

    summary = run_job_search(
        _config(
            sources={
                "france_travail": {"enabled": False},
                "adzuna": {
                    "enabled": True,
                    "country": "fr",
                    "max_pages": 1,
                    "results_per_page": 20,
                    "keywords": ["frontend"],
                    "location": "Dijon",
                },
            }
        ),
        _env(),
        client=client,
        adzuna_client=_adzuna_client([_adzuna_offer("AZ1")], calls=adzuna_calls),
        data_dir=tmp_path / "data",
        export_dir=tmp_path / "exports",
    )

    assert client.calls == []
    assert len(adzuna_calls) == 1
    assert "https://api.adzuna.com/v1/api/jobs/fr/search/1" in adzuna_calls[0]
    assert summary["total_raw"] == 1
    assert summary["total_relevant"] == 1
    assert summary["sources_enabled"] == ["Adzuna"]
    assert summary["source_counts"] == {"Adzuna": {"raw": 1, "normalized": 1}}
    assert summary["source_stats"]["France Travail"]["enabled"] is False
    assert summary["source_stats"]["Arbeitnow"]["enabled"] is False
    assert summary["source_stats"]["Remotive"]["enabled"] is False
    assert summary["source_stats"]["Adzuna"] == {
        "enabled": True,
        "fetched": 1,
        "kept": 1,
        "filtered": 0,
    }


def test_runner_adzuna_enabled_without_credentials_raises_clear_error(tmp_path: Path) -> None:
    client = FakeFranceTravailClient([[_wordpress_offer()]])
    calls = []

    try:
        run_job_search(
            _config(sources={"france_travail": {"enabled": False}, "adzuna": {"enabled": True}}),
            _env(adzuna_app_id="", adzuna_app_key=""),
            client=client,
            adzuna_client=_adzuna_client([_adzuna_offer()], calls=calls),
            data_dir=tmp_path / "data",
            export_dir=tmp_path / "exports",
        )
    except RuntimeError as exc:
        message = str(exc)
    else:
        raise AssertionError("Adzuna activee sans identifiants devrait echouer.")

    assert calls == []
    assert "Identifiants Adzuna manquants" in message
    assert "fake-adzuna" not in message
    assert "ADZUNA_APP_ID" not in message
    assert "ADZUNA_APP_KEY" not in message


def test_runner_merges_all_sources_including_adzuna(tmp_path: Path) -> None:
    client = FakeFranceTravailClient([[_wordpress_offer("FT1")]])

    summary = run_job_search(
        _config(
            sources={
                "france_travail": {"enabled": True},
                "arbeitnow": {"enabled": True, "max_pages": 1},
                "remotive": {"enabled": True, "keywords": ["frontend"]},
                "adzuna": {"enabled": True, "keywords": ["frontend"]},
            }
        ),
        _env(),
        client=client,
        arbeitnow_client=_arbeitnow_client([_arbeitnow_offer("AN1")]),
        remotive_client=_remotive_client([_remotive_offer("RM1")]),
        adzuna_client=_adzuna_client([_adzuna_offer("AZ1")]),
        data_dir=tmp_path / "data",
        export_dir=tmp_path / "exports",
        include_debug_offers=True,
    )

    assert summary["total_raw"] == 4
    assert summary["total_unique_normalized"] == 4
    assert summary["total_relevant"] == 4
    assert summary["sources_enabled"] == ["France Travail", "Arbeitnow", "Remotive", "Adzuna"]
    assert summary["source_counts"] == {
        "France Travail": {"raw": 1, "normalized": 1},
        "Arbeitnow": {"raw": 1, "normalized": 1},
        "Remotive": {"raw": 1, "normalized": 1},
        "Adzuna": {"raw": 1, "normalized": 1},
    }
    assert summary["source_stats"]["Adzuna"] == {
        "enabled": True,
        "fetched": 1,
        "kept": 1,
        "filtered": 0,
    }
    assert {offer["source"] for offer in summary["debug_offers"]} == {
        "France Travail",
        "Arbeitnow",
        "Remotive",
        "Adzuna",
    }


def test_runner_can_use_jooble_when_other_sources_disabled(tmp_path: Path) -> None:
    client = FakeFranceTravailClient([[_wordpress_offer()]])
    jooble_calls = []

    summary = run_job_search(
        _config(
            sources={
                "france_travail": {"enabled": False},
                "jooble": {
                    "enabled": True,
                    "base_url": "https://example.test/jooble-runner",
                    "max_pages": 1,
                    "keywords": ["web"],
                    "location": "Dijon",
                },
            }
        ),
        _env(),
        client=client,
        jooble_client=_jooble_client([_jooble_offer("JB1")], calls=jooble_calls),
        data_dir=tmp_path / "data",
        export_dir=tmp_path / "exports",
    )

    assert client.calls == []
    assert len(jooble_calls) == 1
    assert "https://example.test/jooble-runner/fake-jooble-key" in jooble_calls[0]
    assert summary["total_raw"] == 1
    assert summary["total_relevant"] == 1
    assert summary["sources_enabled"] == ["Jooble"]
    assert summary["source_counts"] == {"Jooble": {"raw": 1, "normalized": 1}}
    assert summary["source_stats"]["France Travail"]["enabled"] is False
    assert summary["source_stats"]["Arbeitnow"]["enabled"] is False
    assert summary["source_stats"]["Remotive"]["enabled"] is False
    assert summary["source_stats"]["Adzuna"]["enabled"] is False
    assert summary["source_stats"]["Jooble"] == {
        "enabled": True,
        "fetched": 1,
        "kept": 1,
        "filtered": 0,
    }


def test_runner_can_use_themuse_when_other_sources_disabled(tmp_path: Path) -> None:
    themuse_calls = []

    summary = run_job_search(
        _config(
            sources={
                "france_travail": {"enabled": False},
                "themuse": {
                    "enabled": True,
                    "base_url": "https://example.test/themuse-runner",
                    "max_pages": 1,
                    "page_size": 20,
                    "keywords": ["wordpress"],
                    "location": "France",
                },
            }
        ),
        _env(),
        data_dir=tmp_path / "data",
        export_dir=tmp_path / "exports",
        themuse_client=_themuse_client([_themuse_offer("TM1")], calls=themuse_calls),
        include_debug_offers=True,
    )

    assert len(themuse_calls) == 1
    assert "https://example.test/themuse-runner" in themuse_calls[0]
    assert "page=1" in themuse_calls[0]
    assert "per_page=20" in themuse_calls[0]
    assert summary["sources_enabled"] == ["The Muse"]
    assert summary["source_counts"] == {"The Muse": {"raw": 1, "normalized": 1}}
    assert summary["source_stats"]["France Travail"]["enabled"] is False
    assert summary["source_stats"]["The Muse"] == {
        "enabled": True,
        "fetched": 1,
        "kept": 1,
        "filtered": 0,
    }
    assert summary["debug_offers"][0]["source"] == "The Muse"


def test_runner_jooble_enabled_without_key_raises_clear_error(tmp_path: Path) -> None:
    client = FakeFranceTravailClient([[_wordpress_offer()]])
    calls = []

    try:
        run_job_search(
            _config(sources={"france_travail": {"enabled": False}, "jooble": {"enabled": True}}),
            _env(jooble_api_key=""),
            client=client,
            jooble_client=_jooble_client([_jooble_offer()], calls=calls),
            data_dir=tmp_path / "data",
            export_dir=tmp_path / "exports",
        )
    except RuntimeError as exc:
        message = str(exc)
    else:
        raise AssertionError("Jooble activee sans cle devrait echouer.")

    assert calls == []
    assert "Cle Jooble manquante" in message
    assert "fake-jooble-key" not in message
    assert "JOOBLE_API_KEY" not in message


def test_runner_merges_all_sources_including_jooble(tmp_path: Path) -> None:
    client = FakeFranceTravailClient([[_wordpress_offer("FT1")]])

    summary = run_job_search(
        _config(
            sources={
                "france_travail": {"enabled": True},
                "arbeitnow": {"enabled": True, "max_pages": 1},
                "remotive": {"enabled": True, "keywords": ["frontend"]},
                "adzuna": {"enabled": True, "keywords": ["frontend"]},
                "jooble": {"enabled": True, "keywords": ["web"]},
            }
        ),
        _env(),
        client=client,
        arbeitnow_client=_arbeitnow_client([_arbeitnow_offer("AN1")]),
        remotive_client=_remotive_client([_remotive_offer("RM1")]),
        adzuna_client=_adzuna_client([_adzuna_offer("AZ1")]),
        jooble_client=_jooble_client([_jooble_offer("JB1")]),
        data_dir=tmp_path / "data",
        export_dir=tmp_path / "exports",
        include_debug_offers=True,
    )

    assert summary["total_raw"] == 5
    assert summary["total_unique_normalized"] == 5
    assert summary["total_relevant"] == 5
    assert summary["sources_enabled"] == ["France Travail", "Arbeitnow", "Remotive", "Adzuna", "Jooble"]
    assert summary["source_counts"] == {
        "France Travail": {"raw": 1, "normalized": 1},
        "Arbeitnow": {"raw": 1, "normalized": 1},
        "Remotive": {"raw": 1, "normalized": 1},
        "Adzuna": {"raw": 1, "normalized": 1},
        "Jooble": {"raw": 1, "normalized": 1},
    }
    assert summary["source_stats"]["Jooble"] == {
        "enabled": True,
        "fetched": 1,
        "kept": 1,
        "filtered": 0,
    }
    assert {offer["source"] for offer in summary["debug_offers"]} == {
        "France Travail",
        "Arbeitnow",
        "Remotive",
        "Adzuna",
        "Jooble",
    }


def test_runner_exposes_arbeitnow_filtered_source_stats(tmp_path: Path) -> None:
    kept = _arbeitnow_offer("kept")
    rejected_keyword = _arbeitnow_offer("keyword-rejected")
    rejected_keyword["title"] = "Sales Manager"
    rejected_keyword["description"] = "Sales partnerships."
    rejected_location = _arbeitnow_offer("location-rejected")
    rejected_location["location"] = "Berlin"

    summary = run_job_search(
        _config(
            sources={
                "france_travail": {"enabled": False},
                "arbeitnow": {
                    "enabled": True,
                    "max_pages": 1,
                    "keywords": ["frontend"],
                    "allowed_locations": ["remote"],
                },
            }
        ),
        _env(),
        client=FakeFranceTravailClient([[]]),
        arbeitnow_client=_arbeitnow_client([kept, rejected_keyword, rejected_location]),
        data_dir=tmp_path / "data",
        export_dir=tmp_path / "exports",
    )

    assert summary["total_raw"] == 1
    assert summary["source_stats"]["Arbeitnow"] == {
        "enabled": True,
        "fetched": 3,
        "kept": 1,
        "filtered": 2,
    }


def test_runner_handles_no_enabled_source_without_crashing(tmp_path: Path) -> None:
    client = FakeFranceTravailClient([[_wordpress_offer()]])

    summary = run_job_search(
        _config(sources={"france_travail": {"enabled": False}, "arbeitnow": {"enabled": False}}),
        _env(),
        client=client,
        arbeitnow_client=_arbeitnow_client([_arbeitnow_offer()]),
        data_dir=tmp_path / "data",
        export_dir=tmp_path / "exports",
        include_debug_offers=True,
    )

    assert client.calls == []
    assert summary["source_status"] == "no_sources_enabled"
    assert summary["sources_enabled"] == []
    assert summary["source_counts"] == {}
    assert summary["source_stats"] == {
        "France Travail": {"enabled": False, "fetched": 0, "kept": 0, "filtered": 0},
        "Arbeitnow": {"enabled": False, "fetched": 0, "kept": 0, "filtered": 0},
        "Remotive": {"enabled": False, "fetched": 0, "kept": 0, "filtered": 0},
        "Adzuna": {"enabled": False, "fetched": 0, "kept": 0, "filtered": 0},
        "Jooble": {"enabled": False, "fetched": 0, "kept": 0, "filtered": 0},
        "The Muse": {"enabled": False, "fetched": 0, "kept": 0, "filtered": 0},
    }
    assert summary["total_raw"] == 0
    assert summary["total_normalized"] == 0
    assert summary["total_relevant"] == 0
    assert summary["total_new"] == 0
    assert summary["export_path"] is None
    assert summary["debug_export_path"] is None
    assert summary["debug_offers"] == []


def test_runner_main_export_excludes_rejected_decision_even_when_relevant(tmp_path: Path, monkeypatch) -> None:
    client = FakeFranceTravailClient([[_wordpress_offer("A1")]])
    monkeypatch.setattr(runner_module, "score_offer", _fake_scoring({"A1": "Rejeté"}))

    summary = run_job_search(
        _config(),
        _env(),
        client=client,
        data_dir=tmp_path / "data",
        export_dir=tmp_path / "exports",
        include_debug_offers=True,
    )

    assert summary["total_relevant"] == 1
    assert summary["total_exportable"] == 0
    assert summary["total_new"] == 0
    assert summary["export_path"] is None
    assert summary["xlsx_export_path"] is None
    assert summary["debug_export_path"] is not None
    assert summary["debug_offers"][0]["id_offre"] == "A1"
    assert summary["debug_offers"][0]["decision"].startswith("Rejet")
    assert not Path(summary["seen_ids_path"]).exists()


def test_runner_main_export_includes_review_and_relevant_decisions(tmp_path: Path, monkeypatch) -> None:
    client = FakeFranceTravailClient(
        [[_wordpress_offer("PERTINENT"), _wordpress_offer("REVIEW"), _wordpress_offer("REJECTED")]]
    )
    monkeypatch.setattr(
        runner_module,
        "score_offer",
        _fake_scoring(
            {
                "PERTINENT": DECISION_RELEVANT,
                "REVIEW": DECISION_REVIEW,
                "REJECTED": "Rejeté",
            }
        ),
    )

    summary = run_job_search(
        _config(),
        _env(),
        client=client,
        data_dir=tmp_path / "data",
        export_dir=tmp_path / "exports",
        include_debug_offers=True,
    )

    assert summary["total_relevant"] == 3
    assert summary["total_exportable"] == 2
    assert summary["total_new"] == 2
    assert summary["export_path"] is not None
    assert summary["xlsx_export_path"] is not None
    with Path(summary["export_path"]).open("r", encoding="utf-8-sig", newline="") as csv_file:
        rows = list(csv.DictReader(csv_file, delimiter=";"))

    assert [row["id_offre"] for row in rows] == ["PERTINENT", "REVIEW"]
    assert [offer["id_offre"] for offer in summary["debug_offers"]] == ["PERTINENT", "REVIEW", "REJECTED"]
    assert json.loads(Path(summary["seen_ids_path"]).read_text(encoding="utf-8")) == ["PERTINENT", "REVIEW"]


def test_runner_discord_summary_uses_really_exported_offer_count(tmp_path: Path, monkeypatch) -> None:
    client = FakeFranceTravailClient([[_wordpress_offer("A1")]])
    calls = []
    monkeypatch.setattr(runner_module, "score_offer", _fake_scoring({"A1": "Rejeté"}))

    summary = run_job_search(
        _config(notifications={"discord_enabled": True, "notify_when_no_results": True}),
        _env(discord_webhook_url="https://discord.test/webhook"),
        client=client,
        data_dir=tmp_path / "data",
        export_dir=tmp_path / "exports",
        discord_sender=lambda webhook_url, payload: calls.append((webhook_url, payload.copy()))
        or {"sent": True, "status": "sent", "error": None},
    )

    assert summary["total_relevant"] == 1
    assert summary["total_new"] == 0
    assert calls[0][1]["total_new"] == 0
