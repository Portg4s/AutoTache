"""Read offers from debug Excel exports without modifying them."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook


OFFER_COLUMNS = [
    "source",
    "titre",
    "entreprise",
    "localisation",
    "description",
    "technologies",
    "type_contrat",
    "decision",
    "score_total",
    "score_reason",
    "url_offre",
]


class OfferReaderError(ValueError):
    """Raised when an offer cannot be read from an Excel export."""


def read_offer_from_xlsx(path: str | Path, row_number: int) -> dict[str, Any]:
    xlsx_path = Path(path)
    if not xlsx_path.exists():
        raise OfferReaderError(f"Fichier Excel introuvable: {xlsx_path}")
    if row_number <= 1:
        raise OfferReaderError("--row doit correspondre a une ligne Excel de donnees (2 ou plus).")

    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    worksheet = workbook.active
    try:
        if row_number > worksheet.max_row:
            raise OfferReaderError(f"Ligne Excel invalide: {row_number} (maximum {worksheet.max_row}).")

        columns = header_columns(worksheet)
        row_values = [cell.value for cell in worksheet[row_number]]
        if not any(_has_value(value) for value in row_values):
            raise OfferReaderError(f"Ligne Excel vide: {row_number}.")

        offer = {column: "" for column in OFFER_COLUMNS}
        for column in OFFER_COLUMNS:
            column_index = columns.get(column)
            if column_index is not None:
                offer[column] = worksheet.cell(row=row_number, column=column_index).value or ""
        offer["row"] = row_number
        return offer
    finally:
        workbook.close()


def read_all_offers_from_xlsx(path: str | Path) -> list[dict[str, Any]]:
    xlsx_path = Path(path)
    if not xlsx_path.exists():
        raise OfferReaderError(f"Fichier Excel introuvable: {xlsx_path}")

    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    worksheet = workbook.active
    try:
        columns = header_columns(worksheet)
        offers: list[dict[str, Any]] = []
        for row_number in range(2, worksheet.max_row + 1):
            row_values = [cell.value for cell in worksheet[row_number]]
            if not any(_has_value(value) for value in row_values):
                continue

            offer = {column: "" for column in OFFER_COLUMNS}
            for column in OFFER_COLUMNS:
                column_index = columns.get(column)
                if column_index is not None:
                    offer[column] = worksheet.cell(row=row_number, column=column_index).value or ""
            offer["row"] = row_number
            offers.append(offer)
        return offers
    finally:
        workbook.close()


def header_columns(worksheet: Any) -> dict[str, int]:
    return {
        str(cell.value).strip(): cell.column
        for cell in worksheet[1]
        if cell.value is not None and str(cell.value).strip()
    }


def _has_value(value: Any) -> bool:
    return value is not None and str(value).strip() != ""

